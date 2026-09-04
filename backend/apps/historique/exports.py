import os

import openpyxl
from django.conf import settings
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def export_fiche_article_horizontal(article_code, historique, filename):
    """
    Exporte la fiche article au format horizontal (dates en colonnes)
    Comme l'Excel montré par l'utilisateur
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Fiche {article_code}"
    
    # Style
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    entree_fill = PatternFill(start_color="86EFAC", end_color="86EFAC", fill_type="solid")
    sortie_fill = PatternFill(start_color="F87171", end_color="F87171", fill_type="solid")
    reste_fill = PatternFill(start_color="FDE68A", end_color="FDE68A", fill_type="solid")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    
    # Groupement par date
    mouvements_par_date = {}
    for mvt in historique:
        date_str = mvt["date"].strftime("%d/%m/%Y")
        if date_str not in mouvements_par_date:
            mouvements_par_date[date_str] = {"entree": 0, "sortie": 0, "reste": 0}
        
        if mvt["impact"] > 0:
            mouvements_par_date[date_str]["entree"] += mvt["impact"]
        elif mvt["impact"] < 0:
            mouvements_par_date[date_str]["sortie"] += abs(mvt["impact"])
        
        mouvements_par_date[date_str]["reste"] = mvt["stock_cumule"]
    
    # Headers
    row = 1
    col = 1
    ws.cell(row=row, column=col, value="ARTICLES").font = header_font
    ws.cell(row=row, column=col).fill = header_fill
    ws.cell(row=row, column=col).border = border
    ws.cell(row=row, column=col+1, value="En stock réel").font = header_font
    ws.cell(row=row, column=col+1).fill = header_fill
    ws.cell(row=row, column=col+1).border = border
    
    # Dates en colonnes
    dates = sorted(mouvements_par_date.keys())
    for i, date_str in enumerate(dates):
        ws.cell(row=row, column=col+2+(i*3), value="Entrée").font = header_font
        ws.cell(row=row, column=col+2+(i*3)).fill = entree_fill
        ws.cell(row=row, column=col+2+(i*3)).border = border
        ws.cell(row=row, column=col+3+(i*3), value="Sortie").font = header_font
        ws.cell(row=row, column=col+3+(i*3)).fill = sortie_fill
        ws.cell(row=row, column=col+3+(i*3)).border = border
        ws.cell(row=row, column=col+4+(i*3), value="Reste").font = header_font
        ws.cell(row=row, column=col+4+(i*3)).fill = reste_fill
        ws.cell(row=row, column=col+4+(i*3)).border = border
    
    # Ligne des dates
    row = 2
    ws.cell(row=row, column=col, value="DATE").font = Font(bold=True)
    ws.cell(row=row, column=col).border = border
    ws.cell(row=row, column=col+1).border = border
    for i, date_str in enumerate(dates):
        ws.cell(row=row, column=col+1+(i*3), value=date_str).border = border
        ws.cell(row=row, column=col+2+(i*3), value=date_str).border = border
        ws.cell(row=row, column=col+3+(i*3), value=date_str).border = border
    
    # Données
    row = 3
    ws.cell(row=row, column=col, value=article_code).border = border
    # Stock initial
    stock_initial = historique[0]["stock_cumule"] if historique else 0
    ws.cell(row=row, column=col+1, value=stock_initial).border = border
    
    # Mouvements par date
    for i, date_str in enumerate(dates):
        data = mouvements_par_date[date_str]
        ws.cell(row=row, column=col+2+(i*3), value=data["entree"]).border = border
        ws.cell(row=row, column=col+3+(i*3), value=data["sortie"]).border = border
        ws.cell(row=row, column=col+4+(i*3), value=data["reste"]).border = border
    
    # Ajuster les colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Sauvegarder dans le dossier media
    filepath = os.path.join(settings.MEDIA_ROOT, "exports", filename)
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)
    
    return filepath